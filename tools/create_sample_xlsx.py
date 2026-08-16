from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


source = Path(r"%USERPROFILE%\Downloads\新灵芝茶拍摄脚本分镜-抗氧化-4.xlsx")
output = Path(r"%USERPROFILE%\Downloads\新灵芝茶拍摄脚本分镜-工具版示例.xlsx")

headers = [
    "脚本",
    "镜头",
    "状态",
    "时长(秒)",
    "景别",
    "运镜",
    "分镜画面",
    "口播稿",
    "分镜图",
    "分镜图链接",
    "视频链接",
]

rows = [
    [
        "开场钩子：朋友，蓝莓、蔓越莓、葡萄籽你都听过，但今天这杯灵芝茶更适合日常抗氧化。",
        "镜头 1",
        "待拍摄",
        6,
        "中景",
        "缓慢推近",
        "主播坐在干净茶桌前，桌面摆放新灵芝茶、透明茶杯和几种常见抗氧化食材；镜头从桌面产品推到主播。",
        "朋友，蓝莓、蔓越莓、葡萄籽你都听过，但今天我想让你看看这杯新灵芝茶。",
        "",
        "",
        "",
    ],
    [
        "产品背书：官方联合华中农业大学共同推出。",
        "镜头 2",
        "待拍摄",
        4,
        "近景",
        "固定",
        "产品包装正面特写，画面干净，手轻扶包装边缘，字幕强调“联合研发”。",
        "它是官方联合华中农业大学共同推出的新灵芝茶。",
        "",
        "",
        "",
    ],
    [
        "核心卖点：一包里灵芝含量超过 30%。",
        "镜头 3",
        "待拍摄",
        4,
        "特写",
        "俯拍推进",
        "撕开茶包或展示原料，白色托盘上能看到灵芝片和茶料，光线从侧前方打出质感。",
        "一包里面，灵芝含量超过百分之三十。",
        "",
        "",
        "",
    ],
    [
        "品质来源：灵芝由华中农业大学技术支持种植。",
        "镜头 4",
        "待拍摄",
        5,
        "特写",
        "横移",
        "包装上的产地、研发或品质信息标签细节，手指轻轻指向关键文字，画面稳。",
        "灵芝的种植和品质把控，也有专业技术支持。",
        "",
        "",
        "",
    ],
    [
        "使用场景：细胞怕氧化，日常饮食也要更轻负担。",
        "镜头 5",
        "待拍摄",
        5,
        "中近景",
        "切换硬切",
        "画面切到办公桌、熬夜电脑、下午茶等生活场景，穿插主播拿起茶杯。",
        "平时忙、熬夜、饮食不规律的时候，日常抗氧化就更要跟上。",
        "",
        "",
        "",
    ],
    [
        "冲泡过程：热水冲泡，茶汤颜色自然清透。",
        "镜头 6",
        "待拍摄",
        4,
        "特写",
        "慢动作",
        "热水倒入玻璃杯，茶汤颜色慢慢舒展开，蒸汽轻微上升，突出自然感。",
        "",
        "",
        "",
        "",
    ],
    [
        "饮用体验：口感清润，适合日常喝。",
        "镜头 7",
        "待拍摄",
        4,
        "近景",
        "固定",
        "主播端杯轻抿一口，表情自然放松，杯子和包装都在画面里。",
        "口感是比较清润的，不会有很重的负担感，日常喝刚刚好。",
        "",
        "",
        "",
    ],
    [
        "配方画面：灵芝、枸杞、玫瑰、桑葚等食材展示。",
        "镜头 8",
        "待拍摄",
        5,
        "俯拍全景",
        "平移扫过",
        "白色桌面整齐摆放配方食材，镜头从左到右扫过，每个食材旁边有小标签。",
        "里面还搭配了枸杞、玫瑰、桑葚这些日常熟悉的食材。",
        "",
        "",
        "",
    ],
    [
        "人群说明：适合想做日常养护、关注状态的人。",
        "镜头 9",
        "待拍摄",
        5,
        "中景",
        "轻微推近",
        "主播面向镜头，手边放茶杯，画面叠加“日常抗氧化 / 清润茶饮 / 办公室也能喝”。",
        "如果你平时也想把抗氧化这件事做得更简单，可以把它放进每天的茶饮里。",
        "",
        "",
        "",
    ],
    [
        "结尾转化：官方推出，价格更友好，下单带走。",
        "镜头 10",
        "待拍摄",
        5,
        "近景",
        "固定后切产品特写",
        "主播把产品推到镜头前，最后切到产品包装和冲好的茶汤同框，画面留出价格/权益字幕位置。",
        "现在官方推出，价格也很友好。想试试新灵芝茶的朋友，可以直接下单带走。",
        "",
        "",
        "",
    ],
]


def clone_basic_style(ws):
    if not source.exists():
        return
    src_wb = load_workbook(source)
    src_ws = src_wb.active
    for row in range(1, min(3, src_ws.max_row) + 1):
        for col in range(1, min(len(headers), src_ws.max_column) + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = ws.cell(row=row, column=col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
    if src_ws.row_dimensions[1].height:
        ws.row_dimensions[1].height = src_ws.row_dimensions[1].height


wb = Workbook()
ws = wb.active
ws.title = "分镜执行表"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
ws["A1"] = "新灵芝茶拍摄脚本分镜 - 抗氧化工具版示例"
ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="5B9BD5")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, row in enumerate(rows, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="微软雅黑", size=10)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx in (2, 3, 4, 5, 6):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 68

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.border = border

widths = {
    "A": 28,
    "B": 10,
    "C": 12,
    "D": 10,
    "E": 12,
    "F": 12,
    "G": 42,
    "H": 38,
    "I": 18,
    "J": 24,
    "K": 24,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:K{ws.max_row}"

dv = DataValidation(type="list", formula1='"待拍摄,已拍摄,需重拍,已剪辑,已完成"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"C3:C{ws.max_row}")

note_ws = wb.create_sheet("拍摄清单")
note_ws["A1"] = "拍摄顺序建议"
note_ws["A1"].font = Font(name="微软雅黑", size=16, bold=True)
checklist_rows = [
    ("1", "空镜/产品细节", "先拍产品包装、茶汤、原料、桌面环境，不需要主播等待。"),
    ("2", "B-roll 补充", "拍办公桌、下午茶、冲泡过程、食材扫拍，用来遮挡口播剪辑。"),
    ("3", "口播录制", "主播集中录制镜头 1、2、5、7、9、10，统一妆发和情绪。"),
    ("4", "道具", "新灵芝茶产品、透明茶杯、热水壶、白色托盘、配方食材小碟、标签卡。"),
]
for idx, row in enumerate(checklist_rows, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = note_ws.cell(row=idx, column=col_idx, value=value)
        cell.font = Font(name="微软雅黑", size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
note_ws.column_dimensions["A"].width = 8
note_ws.column_dimensions["B"].width = 18
note_ws.column_dimensions["C"].width = 80
for row_idx in range(3, 7):
    note_ws.row_dimensions[row_idx].height = 42

output.parent.mkdir(parents=True, exist_ok=True)
wb.save(output)
print(output)
