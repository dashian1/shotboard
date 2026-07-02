from openpyxl import load_workbook

path = r"C:\Users\gba\Downloads\新灵芝茶拍摄脚本分镜-抗氧化-4.xlsx"
wb = load_workbook(path)
print(wb.sheetnames)
ws = wb.active
print(ws.title, ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
    print(list(row))
print("widths", [(cell.column_letter, ws.column_dimensions[cell.column_letter].width) for cell in ws[1]])
